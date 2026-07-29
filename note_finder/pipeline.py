from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook

from .dart import DartClient, DartError
from .extract import extract_document

FIELDS = ["rcept_dt", "rcept_no", "corp_name", "corp_code", "report_nm", "source_file",
          "context", "broker", "raw_amount", "raw_unit", "amount_thousand_won",
          "classification", "exposure_type", "reason", "dart_url"]

REPORT_PERIOD = re.compile(r"\((\d{4}\.\d{2})\)")
SUPPORTING_REPORT = re.compile(r"(?:감사|검토)보고서$")


def dart_url(rcept_no: str, keyword: str = "발행어음") -> str:
    return (
        "https://dart.fss.or.kr/dsaf001/main.do"
        f"?rcpNo={rcept_no}&keyword={quote(keyword)}"
    )


def unique_receipts(filings: list[dict]) -> tuple[list[dict], list[dict]]:
    """Collapse repeated DART search rows without conflating them with companies."""
    kept: dict[str, dict] = {}
    duplicate_log = []
    for filing in filings:
        rcept_no = str(filing.get("rcept_no", "")).strip()
        if not rcept_no:
            continue
        if rcept_no in kept:
            duplicate_log.append({
                "rcept_no": rcept_no,
                "corp_name": filing.get("corp_name", ""),
            })
            continue
        kept[rcept_no] = filing
    return list(kept.values()), duplicate_log


def _company_key(filing: dict) -> str:
    return str(
        filing.get("corp_code")
        or filing.get("corp_name")
        or filing.get("rcept_no")
        or ""
    ).strip()


def _latest_key(filing: dict) -> tuple[str, int, str, str]:
    report = filing.get("report_nm", "")
    period = REPORT_PERIOD.search(report)
    return (
        period.group(1) if period else "",
        0 if SUPPORTING_REPORT.search(report) else 1,
        filing.get("rcept_dt", ""),
        filing.get("rcept_no", ""),
    )


def select_latest_by_company(filings: list[dict]) -> list[dict]:
    """Select one current periodic disclosure per corporation for final review."""
    grouped: dict[str, list[dict]] = defaultdict(list)
    for filing in filings:
        grouped[_company_key(filing)].append(filing)
    return [
        max(group, key=_latest_key)
        for _, group in sorted(grouped.items())
    ]


def deduplicate(filings: list[dict]) -> tuple[list[dict], list[dict]]:
    kept, removed = {}, []
    for filing in sorted(filings, key=lambda x: (x.get("rcept_dt", ""), x["rcept_no"])):
        report = filing.get("report_nm", "").replace("[기재정정]", "").strip()
        key = (filing.get("corp_code") or filing.get("corp_name"), report)
        if key in kept:
            removed.append({"dropped": kept[key]["rcept_no"], "kept": filing["rcept_no"], "key": str(key)})
        kept[key] = filing
    return list(kept.values()), removed


def correction_fallbacks(filings: list[dict]) -> dict[str, list[dict]]:
    groups: dict[tuple, list[dict]] = {}
    for filing in sorted(filings, key=lambda x: (x.get("rcept_dt", ""), x["rcept_no"])):
        report = filing.get("report_nm", "").replace("[기재정정]", "").strip()
        key = (filing.get("corp_code") or filing.get("corp_name"), report)
        groups.setdefault(key, []).append(filing)
    return {
        group[-1]["rcept_no"]: list(reversed(group[:-1]))
        for group in groups.values()
        if len(group) > 1
    }


def run(client: DartClient, begin: str, end: str, output: Path, keyword: str = "발행어음",
        candidate_filings: list[dict] | None = None, latest_per_company: bool = True,
        search_result_count: int | None = None) -> dict:
    raw = candidate_filings if candidate_filings is not None else list(client.iter_filings(begin, end))
    receipts, duplicate_receipt_log = unique_receipts(raw)
    corrected, dedup_log = deduplicate(receipts)
    filings = select_latest_by_company(corrected) if latest_per_company else corrected
    fallbacks = correction_fallbacks(receipts)
    rows, document_errors, correction_fallback_log = [], [], []
    for filing in filings:
        selected, documents = filing, None
        for choice in [filing, *fallbacks.get(filing["rcept_no"], [])]:
            try:
                documents = client.document(choice["rcept_no"])
                selected = choice
                if choice is not filing:
                    correction_fallback_log.append({
                        "unavailable": filing["rcept_no"],
                        "used": choice["rcept_no"],
                    })
                break
            except DartError as exc:
                document_errors.append({"rcept_no": choice["rcept_no"], "error": str(exc)})
        if documents is None:
            continue
        for name, body in documents:
            for evidence in extract_document(name, body, keyword):
                rows.append({
                    **{k: selected.get(k, "") for k in FIELDS[:5]},
                    **evidence.dict(),
                    "dart_url": dart_url(selected["rcept_no"], keyword),
                })
    audit_summary = {
        "search_rows": search_result_count if search_result_count is not None else len(raw),
        "candidate_rows": len(raw),
        "unique_receipts": len(receipts),
        "unique_companies": len({_company_key(filing) for filing in receipts}),
        "filings_after_correction_dedup": len(corrected),
        "filings_scanned": len(filings),
        "evidence_rows": len(rows),
        "latest_per_company": latest_per_company,
    }
    write_excel(output, rows, dedup_log, audit_summary, duplicate_receipt_log)
    audit = output.with_suffix(".audit.json")
    audit.write_text(json.dumps({"query": {"begin": begin, "end": end, "keyword": keyword,
                                           "source": "candidate_file" if candidate_filings is not None else "opendart_list"},
                                 **audit_summary, "dedup_log": dedup_log,
                                 "duplicate_receipt_log": duplicate_receipt_log,
                                 "document_errors": document_errors,
                                 "correction_fallback_log": correction_fallback_log},
                                ensure_ascii=False, indent=2), encoding="utf-8")
    return {**audit_summary, "rows": len(rows), "document_errors": len(document_errors),
            "output": str(output), "audit": str(audit)}


def write_excel(path: Path, rows: list[dict], dedup_log: list[dict],
                audit_summary: dict | None = None,
                duplicate_receipt_log: list[dict] | None = None) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("audit_summary")
    ws.append(["metric", "value"])
    for key, value in (audit_summary or {}).items():
        ws.append([key, value])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    groups = {"confirmed": "A", "needs_review": None, "excluded": "EXCLUDED"}
    for sheet_name, classification in groups.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(FIELDS)
        selected = [r for r in rows if (r["classification"] == classification if classification else r["classification"] in {"B", "REVIEW"})]
        for row in selected:
            ws.append([row.get(field) for field in FIELDS])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    ws = wb.create_sheet("dedup_log")
    ws.append(["dropped", "kept", "key"])
    for row in dedup_log:
        ws.append([row["dropped"], row["kept"], row["key"]])
    ws = wb.create_sheet("duplicate_receipts")
    ws.append(["rcept_no", "corp_name"])
    for row in duplicate_receipt_log or []:
        ws.append([row["rcept_no"], row["corp_name"]])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
