from __future__ import annotations

from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


TYPE_ORDER = {
    "DIRECT": 1,
    "AGGREGATE": 2,
    "INDIRECT": 3,
    "TRUST": 4,
    "EXISTENCE": 5,
}


def keyword_url(url: str, keyword: str = "발행어음") -> str:
    if "keyword=" in url:
        return url
    separator = "&" if "?" in url else "?"
    return f"{url}{separator}keyword={quote(keyword)}"


def export_reviewed_workbook(
    curated: dict,
    output: Path,
    *,
    candidate_rows: int,
    unique_receipts: int,
    search_result_count: int | None = None,
) -> None:
    """Export the reviewed company list in the final five-column handoff format."""
    rows = list(curated["company_rows"])
    rows.sort(
        key=lambda row: (
            min(
                (
                    TYPE_ORDER.get(value.strip(), 99)
                    for value in row.get("evidence_type", "").split(",")
                ),
                default=99,
            ),
            -(row["amount_won"] if row["amount_won"] is not None else -1),
            row["corp_name"],
        )
    )
    excluded_count = len(curated.get("excluded_rows", []))
    company_count = len(rows) + excluded_count

    wb = Workbook()
    ws = wb.active
    ws.title = f"발행어음 {len(rows)}개사"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B5"

    navy = "17365D"
    blue = "1F4E78"
    gray = "F2F2F2"
    violet = "E4DFEC"

    ws.merge_cells("A1:E1")
    ws["A1"] = "최근 1년 DART 발행어음 확인 기업"
    ws["A1"].fill = PatternFill("solid", fgColor=navy)
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:E2")
    ws["A2"] = (
        f"총 {len(rows)}개사  |  보유·간접·신탁 내역 포함  |  "
        "자사 발행부채 제외  |  빈 금액은 공시 미표기"
    )
    ws["A2"].fill = PatternFill("solid", fgColor=gray)
    ws["A2"].font = Font(color="595959", size=10)
    ws.row_dimensions[2].height = 28

    ws.merge_cells("A3:E3")
    prefix = (
        f"DART 검색 {search_result_count}행"
        if search_result_count is not None
        else f"후보 CSV {candidate_rows}행"
    )
    ws["A3"] = (
        f"전수 재검토: {prefix} → 고유 접수번호 {unique_receipts}건 "
        f"→ 회사 {company_count}개 → 포함 {len(rows)}개사 / 제외 {excluded_count}개사"
    )
    ws["A3"].fill = PatternFill("solid", fgColor=violet)
    ws["A3"].font = Font(color="60497A", bold=True, size=10)
    ws.row_dimensions[3].height = 24

    headers = ["회사명", "확인금액(원)", "발행 증권사", "공시상 표현", "DART 원문"]
    ws.append(headers)
    for cell in ws[4]:
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 30

    for row in rows:
        source_url = keyword_url(row["dart_url"])
        ws.append([
            row["corp_name"],
            row["amount_won"],
            row.get("issuer") or "공시 내 미표기",
            row["expression"],
            source_url,
        ])
        excel_row = ws.max_row
        ws.cell(excel_row, 2).number_format = "#,##0;[Red](#,##0);-"
        ws.cell(excel_row, 2).alignment = Alignment(horizontal="right")
        ws.cell(excel_row, 3).alignment = Alignment(horizontal="center")
        ws.cell(excel_row, 4).alignment = Alignment(wrap_text=True, vertical="center")
        link = ws.cell(excel_row, 5)
        link.hyperlink = source_url
        link.style = "Hyperlink"
        link.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[excel_row].height = 36
        if row["amount_won"] is None:
            ws.cell(excel_row, 2).fill = PatternFill("solid", fgColor=gray)

    end_row = ws.max_row
    table = Table(displayName="IssuedNoteReviewedCompanies", ref=f"A4:E{end_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    widths = {"A": 21, "B": 20, "C": 24, "D": 76, "E": 58}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
