import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from note_finder.extract import extract_document, find_brokers
from note_finder.pipeline import (
    dart_url,
    deduplicate,
    run,
    select_latest_by_company,
    unique_receipts,
    write_excel,
)
from note_finder.review import export_reviewed_workbook, keyword_url


class HarnessTests(unittest.TestCase):
    def test_matches_any_explicit_securities_company_not_only_four_aliases(self):
        self.assertEqual(find_brokers("삼성증권 발행어음 1,000천원"), ["삼성증권"])
        self.assertEqual(find_brokers("키움증권 및 NH투자증권"), ["키움증권", "NH투자증권"])

    def test_preserves_disclosed_spelling_instead_of_forcing_aliases(self):
        self.assertEqual(find_brokers("케이비증권 및 KB증권"), ["케이비증권", "KB증권"])

    def test_classifies_explicit_and_aggregate_without_dropping(self):
        html = """<table><tr><td>NH투자증권</td><td>발행어음</td><td>11,200 천원</td></tr>
        <tr><td>기타금융상품(중금채, 발행어음 등)</td><td>35,960,207 천원</td></tr></table>"""
        rows = extract_document("sample.xml", html.encode())
        self.assertEqual([row.classification for row in rows], ["A", "B"])
        self.assertEqual(rows[0].amount_thousand_won, 11200)

    def test_confirms_holding_when_issuer_is_not_disclosed(self):
        html = "<table><tr><td>단기금융상품</td><td>발행어음</td><td>60,000</td></tr></table>"
        row = extract_document("sample.xml", html.encode())[0]
        self.assertEqual(row.classification, "A")
        self.assertEqual(row.broker, "공시 내 미표기")

    def test_excludes_unissued_note_false_positive(self):
        html = "<p>은행으로부터 교부받은 미발행어음 26매가 있습니다.</p>"
        row = extract_document("sample.xml", html.encode())[0]
        self.assertEqual(row.classification, "EXCLUDED")
        self.assertEqual(row.exposure_type, "NOT_ISSUED")

    def test_excludes_issuer_liability_instead_of_treating_it_as_holding(self):
        html = "<table><tr><td>발행어음예수부채</td><td>7,624,245 천원</td></tr></table>"
        row = extract_document("sample.xml", html.encode())[0]
        self.assertEqual(row.classification, "EXCLUDED")
        self.assertEqual(row.exposure_type, "LIABILITY")

    def test_keeps_trust_and_indirect_exposure_for_review(self):
        html = """<table>
        <tr><td>신탁계정 매입어음 - 발행어음</td><td>264,670 백만원</td></tr>
        <tr><td>MMT 구성자산은 발행어음과 RP</td><td>2,273,279 천원</td></tr>
        </table>"""
        rows = extract_document("sample.xml", html.encode())
        self.assertEqual(
            [(row.classification, row.exposure_type) for row in rows],
            [("A", "TRUST"), ("B", "INDIRECT")],
        )

    def test_does_not_treat_direct_operating_amount_as_indirect(self):
        html = (
            "<table><tr><td>발행어음(외화) USD</td>"
            "<td>운용금액</td><td>1,403 백만원</td></tr></table>"
        )
        row = extract_document("sample.xml", html.encode())[0]
        self.assertEqual(row.classification, "A")
        self.assertEqual(row.exposure_type, "DIRECT")

    def test_keeps_existence_only_trust_disclosure_without_inventing_amount(self):
        html = "<p>신탁계정 발행어음이자가 발생하였습니다.</p>"
        row = extract_document("sample.xml", html.encode())[0]
        self.assertEqual(row.classification, "REVIEW")
        self.assertEqual(row.exposure_type, "TRUST")
        self.assertIsNone(row.amount_thousand_won)

    def test_keeps_latest_correction_and_logs_old_receipt(self):
        filings = [{"corp_code": "1", "report_nm": "사업보고서 (2025.12)", "rcept_dt": "20260301", "rcept_no": "old"},
                   {"corp_code": "1", "report_nm": "[기재정정] 사업보고서 (2025.12)", "rcept_dt": "20260401", "rcept_no": "new"}]
        kept, log = deduplicate(filings)
        self.assertEqual(kept[0]["rcept_no"], "new")
        self.assertEqual(log[0]["dropped"], "old")

    def test_distinguishes_search_row_duplicates_from_distinct_filings(self):
        filings = [
            {"rcept_no": "1", "corp_code": "A", "corp_name": "회사A"},
            {"rcept_no": "1", "corp_code": "A", "corp_name": "회사A"},
            {"rcept_no": "2", "corp_code": "A", "corp_name": "회사A"},
        ]
        kept, log = unique_receipts(filings)
        self.assertEqual([row["rcept_no"] for row in kept], ["1", "2"])
        self.assertEqual(len(log), 1)

    def test_selects_latest_main_report_per_company(self):
        filings = [
            {"rcept_no": "1", "rcept_dt": "20260301", "corp_code": "A",
             "report_nm": "사업보고서 (2025.12)"},
            {"rcept_no": "2", "rcept_dt": "20260302", "corp_code": "A",
             "report_nm": "사업보고서 (2025.12) 감사보고서"},
            {"rcept_no": "3", "rcept_dt": "20260515", "corp_code": "A",
             "report_nm": "분기보고서 (2026.03)"},
            {"rcept_no": "4", "rcept_dt": "20260320", "corp_code": "B",
             "report_nm": "사업보고서 (2025.12)"},
        ]
        selected = select_latest_by_company(filings)
        self.assertEqual(
            {row["corp_code"]: row["rcept_no"] for row in selected},
            {"A": "3", "B": "4"},
        )

    def test_builds_dart_link_with_keyword_search(self):
        url = dart_url("20260514000988")
        self.assertIn("rcpNo=20260514000988", url)
        self.assertIn("keyword=%EB%B0%9C%ED%96%89%EC%96%B4%EC%9D%8C", url)
        self.assertEqual(keyword_url(url), url)

    def test_exports_reviewed_companies_without_internal_category_column(self):
        curated = {
            "company_rows": [{
                "corp_name": "예시회사",
                "amount_won": 1_000_000_000,
                "issuer": "공시 내 미표기",
                "expression": "발행어음 | 1,000백만원",
                "dart_url": "https://dart.fss.or.kr/dsaf001/main.do?rcpNo=1",
                "evidence_type": "DIRECT",
            }],
            "excluded_rows": [{"corp_name": "제외회사"}],
        }
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "final.xlsx"
            export_reviewed_workbook(
                curated,
                path,
                candidate_rows=3,
                unique_receipts=2,
                search_result_count=4,
            )
            wb = load_workbook(path)
            ws = wb.active
            self.assertEqual(
                [cell.value for cell in ws[4]],
                ["회사명", "확인금액(원)", "발행 증권사", "공시상 표현", "DART 원문"],
            )
            self.assertNotIn("구분", [cell.value for cell in ws[4]])
            self.assertIn("고유 접수번호 2건", ws["A3"].value)
            self.assertIn("keyword=", ws["E5"].value)

    def test_writes_auditable_workbook(self):
        row = dict.fromkeys(__import__("note_finder.pipeline", fromlist=["FIELDS"]).FIELDS, "")
        row.update(classification="A", context="NH투자증권 발행어음 11,200천원")
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "out.xlsx"
            write_excel(path, [row], [], {"search_rows": 2, "unique_receipts": 1}, [])
            wb = load_workbook(path)
            self.assertEqual(
                wb.sheetnames,
                [
                    "audit_summary",
                    "confirmed",
                    "needs_review",
                    "excluded",
                    "dedup_log",
                    "duplicate_receipts",
                ],
            )
            self.assertEqual(wb["audit_summary"]["B2"].value, 2)
            self.assertEqual(wb["confirmed"].max_row, 2)

    def test_pipeline_audits_search_rows_receipts_and_companies_separately(self):
        class FakeClient:
            def document(self, rcept_no):
                return [(
                    f"{rcept_no}.xml",
                    "<table><tr><td>단기금융상품</td><td>발행어음</td>"
                    "<td>1,000 천원</td></tr></table>".encode(),
                )]

        candidates = [
            {"rcept_no": "1", "rcept_dt": "20260301", "corp_code": "A",
             "corp_name": "회사A", "report_nm": "사업보고서 (2025.12)"},
            {"rcept_no": "1", "rcept_dt": "20260301", "corp_code": "A",
             "corp_name": "회사A", "report_nm": "사업보고서 (2025.12)"},
            {"rcept_no": "2", "rcept_dt": "20260515", "corp_code": "A",
             "corp_name": "회사A", "report_nm": "분기보고서 (2026.03)"},
            {"rcept_no": "3", "rcept_dt": "20260515", "corp_code": "B",
             "corp_name": "회사B", "report_nm": "분기보고서 (2026.03)"},
        ]
        with tempfile.TemporaryDirectory() as temp:
            output = Path(temp) / "audit.xlsx"
            result = run(
                FakeClient(),
                "",
                "",
                output,
                candidate_filings=candidates,
                search_result_count=6,
            )
            audit = json.loads(output.with_suffix(".audit.json").read_text())
            self.assertEqual(result["search_rows"], 6)
            self.assertEqual(audit["candidate_rows"], 4)
            self.assertEqual(audit["unique_receipts"], 3)
            self.assertEqual(audit["unique_companies"], 2)
            self.assertEqual(audit["filings_scanned"], 2)


if __name__ == "__main__":
    unittest.main()
